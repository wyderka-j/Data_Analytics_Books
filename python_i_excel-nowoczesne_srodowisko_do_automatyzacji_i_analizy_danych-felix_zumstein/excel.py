"""Ten moduł zapewnia funkcję odczytu i zapisu plików Excela
pozwalającą pobierać i zapisywać listy dwuwymiarowe.
"""
import re
import itertools
import datetime as dt

# Zależności opcjonalne
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pyxlsb
except ImportError:
    pyxlsb = None
try:
    import xlrd
    from xlrd.biffh import error_text_from_code
except ImportError:
    xlrd = None
try:
    import xlwt
except ImportError:
    xlwt = None
try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None


def read(sheet, first_cell="A1", last_cell=None):
    """Wczytanie dwuwymiarowej listy z zakresu Excela.

    Parametery
    ----------
    sheet : obiekt
        Obiekt arkusza xlrd, openpyxl lub pyxlsb
    first_cell : łańcuch znaków lub krotka, parametr opcjonalny
        Lewy górny róg zakresu Excela, który chcesz odczytać.
        Może to być ciąg znaków typu "A1" lub krotka typu wiersz/kolumna (1, 1),
		domyślnie "A1".
    last_cell : łańcuch znaków lub krotka, parametr opcjonalny
        Prawy dolny róg zakresu Excela, który chcesz odczytać.
        Może to być ciąg znaków typu "A1" lub krotka wiersz/kolumna typu (1, 1),
		domyślnie jest to prawa dolna komórka używanego zakresu.

    Zwraca
    -------
    listę
        Dwuwymiarowa lista zawierająca wartości zakresu Excela
    """
    # xlrd
    if xlrd and isinstance(sheet, xlrd.sheet.Sheet):
        # isinstance zwraca True jeśli arkusz jest typu xlrd.sheet.Sheet
        if last_cell is None:
            # rzeczywisty zakres z danymi, nie używany zakres
            last_cell = (sheet.nrows, sheet.ncols)
        # Przekształcenie notacji "A1" w krotki o indeksach rozpoczynających się od 1
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        values = []
        for r in range(first_cell[0] - 1, last_cell[0]):
            row = []
            for c in range(first_cell[1] - 1, last_cell[1]):
                # Obsługa różnych typów komórek
                if sheet.cell(r, c).ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        sheet.cell(r, c).value, sheet.book.datemode)
                elif sheet.cell(r, c).ctype in [xlrd.XL_CELL_EMPTY,
                                                xlrd.XL_CELL_BLANK]:
                    value = None
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_ERROR:
                    value = error_text_from_code[sheet.cell(r, c).value]
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(sheet.cell(r, c).value)
                else:
                    value = sheet.cell(r, c).value
                row.append(value)
            values.append(row)
        return values

    # openpyxl
    elif openpyxl and isinstance(
            sheet,
            (openpyxl.worksheet.worksheet.Worksheet,
             openpyxl.worksheet._read_only.ReadOnlyWorksheet)):
        if last_cell is None:
            # używany zakres
            last_cell = (sheet.max_row, sheet.max_column)
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.cell.coordinate_to_tuple(first_cell)
        if not isinstance(last_cell, tuple):
            last_cell = openpyxl.utils.cell.coordinate_to_tuple(last_cell)
        data = []
        for row in sheet.iter_rows(min_row=first_cell[0], min_col=first_cell[1],
                                   max_row=last_cell[0], max_col=last_cell[1],
                                   values_only=True):
            data.append(list(row))
        return data

    # pyxlsb
    elif pyxlsb and isinstance(sheet, pyxlsb.worksheet.Worksheet):
        errors = {"0x0": "#NULL!", "0x7": "#DIV/0!", "0xf": "#VALUE!",
                  "0x17": "#REF!", "0x1d": "#NAME?", "0x24": "#NUM!",
                  "0x2a": "#N/A"}
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if last_cell and not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        data = []
        # sheet.rows() jest generatorem, który wymaga islice do wycinania
        for row in itertools.islice(sheet.rows(),
                                    first_cell[0] - 1,
                                    last_cell[0] if last_cell else None):
            data.append([errors.get(cell.v, cell.v) for cell in row]
                        [first_cell[1] - 1 : last_cell[1] if last_cell else None])
        return data
    else:
        raise TypeError(f"Couldn't handle sheet of type {type(sheet)}")


def write(sheet, values, first_cell="A1", date_format=None):
    """Zapis dwuwymiarowej listy do zakresu Excela.

    Parametery
    ----------
    sheet : obiekt
        Obiekt arkusza openpyxl, xlsxwriter lub xlwt. Tryb openpyxl 
        write_only=True nie jest obsługiwany.
    values : lista
        Dwuwymiarowa lista wartości
    first_cell : łańcuch znaków lub krotka, parametr opcjonalny
        Lewy górny róg zakresu Excela, gdzie chcesz zapisać obiekt DataFrame.
        Może to być ciąg znaków typu "A1" lub krotka typu wiersz/kolumna (1, 1),
		domyślnie "A1".		
    date_format : łańcuch znaków, parametr opcjonalny
        Akceptowany tylko jeśli arkusz jest arkuszem openpyxl lub xlwt.
		Domyślnie tworzy daty w formacie "mm/dd/rr". W przypadku xlsxwriter
		ustaw format podczas tworzenia skoroszytu podając parametr:
		options={"default_date_format": "mm/dd/rr"}
    """
    # openpyxl
    if openpyxl and isinstance(
            sheet, openpyxl.worksheet.worksheet.Worksheet):
        if date_format is None:
                date_format = "mm/dd/yy"
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.coordinate_to_tuple(first_cell)
        for i, row in enumerate(values):
            for j, value in enumerate(row):
                cell = sheet.cell(row=first_cell[0] + i,
                                  column=first_cell[1] + j)
                cell.value = value
                if date_format and isinstance(value, (dt.datetime, dt.date)):
                    cell.number_format = date_format

    # XlsxWriter
    elif xlsxwriter and isinstance(sheet, xlsxwriter.worksheet.Worksheet):
        if date_format is not None:
            raise ValueError("parametr date_format musi być ustawiony jako opcja skoroszytu")
        if isinstance(first_cell, tuple):
            first_cell = first_cell[0] - 1, first_cell[1] - 1
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for r, row_data in enumerate(values):
            sheet.write_row(first_cell[0] + r, first_cell[1], row_data)

    # xlwt
    elif xlwt and isinstance(sheet, xlwt.Worksheet):
        if date_format is None:
            date_format = "mm/dd/yy"
        date_format = xlwt.easyxf(num_format_str=date_format)
        if isinstance(first_cell, tuple):
            first_cell = (first_cell[0] - 1, first_cell[1] - 1)
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for i, row in enumerate(values):
            for j, cell in enumerate(row):
                if isinstance(cell, (dt.datetime, dt.date)):
                    sheet.write(i + first_cell[0], j + first_cell[1],
                                cell, date_format)
                else:
                    sheet.write(i + first_cell[0], j + first_cell[1],
                                cell)
    else:
        raise TypeError(f"Couldn't handle sheet of type {type(sheet)}")


def xl_cell_to_rowcol(cell_str):
    """
    Konwertuje odwołanie do komórki z notacji A1 na wiersz i kolumnę z indeksem
	liczonym od zera.

    Argumenty:
       cell_str:  łańcuch znaków typu A1

    Zwraca:
        row, col: indeksy wiersza i kolumny komórki liczone od zera

    This function is from XlsxWriter
    Copyright (c) 2013-2020, John McNamara <jmcnamara@cpan.org>
    All rights reserved.

    Redistribution and use in source and binary forms, with or without
    modification, are permitted provided that the following conditions are met:

    1. Redistributions of source code must retain the above copyright notice, this
       list of conditions and the following disclaimer.
    2. Redistributions in binary form must reproduce the above copyright notice,
       this list of conditions and the following disclaimer in the documentation
       and/or other materials provided with the distribution.

    THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND
    ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED
    WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
    DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT OWNER OR CONTRIBUTORS BE LIABLE FOR
    ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES
    (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES;
    LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND
    ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
    (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS
    SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

    The views and conclusions contained in the software and documentation are those
    of the authors and should not be interpreted as representing official policies,
    either expressed or implied, of the FreeBSD Project.

    """
    if not cell_str:
        return 0, 0

    match = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)").match(cell_str)
    col_str = match.group(2)
    row_str = match.group(4)

    # Konwersja łańcucha kolumny z postaci alfabetycznej (base26) na liczbową.
    expn = 0
    col = 0
    for char in reversed(col_str):
        col += (ord(char) - ord("A") + 1) * (26 ** expn)
        expn += 1

    # Konwersja indeksu liczonego od 1 na indeks liczony od zera
    row = int(row_str) - 1
    col -= 1

    return row, col
