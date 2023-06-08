from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


# Katalog zawierający ten plik
this_dir = Path(__file__).resolve().parent

# Wczytanie wszystkich plików
parts = []
for path in (this_dir / "sales_data").rglob("*.xls*"):
    print(f'Reading {path.name}')
    part = pd.read_excel(path)
    parts.append(part)

# Połączenie obiektów DataFrame z każdego pliku w pojedynczy DataFrame
df = pd.concat(parts)

# Przestawienie każdego sklepu w kolumnę i zsumowanie wszystkich transakcji według dat
pivot = pd.pivot_table(df,
                       index="data_transakcji", columns="sklep",
                       values="kwota", aggfunc="sum")

# Resampling na koniec miesiąca i przypisanie nazwy indeksu 
summary = pivot.resample("M").sum()
summary.index.name = "Miesiąc"

# Sortowanie kolumn według przychodu ogółem
summary = summary.loc[:, summary.sum().sort_values().index]

# Dodanie wiersza i kolumny z podsumowaniem: Użycie "append" wraz z "rename"
# jest wygodnym sposobem dodania wiersza do dolnej części DataFrame.
summary.loc[:, "Razem"] = summary.sum(axis=1)
summary = summary.append(summary.sum(axis=0).rename("Razem"))

#### Zapisanie raportu zbiorczego do pliku Excela ####

# Pozycja i liczba wierszy/kolumn DataFrame.
# openpyxl stosuje indeksy liczone od 1.
startrow, startcol = 3, 2
nrows, ncols = summary.shape

with pd.ExcelWriter(this_dir / "sales_report_openpyxl.xlsx",
                    engine="openpyxl", write_only=True) as writer:
    # pandas stosuje indeksy liczone od zera
    summary.to_excel(writer, sheet_name="Arkusz1",
                     startrow=startrow - 1, startcol=startcol - 1)

    # Pobranie skoroszytu openpyxl i abiektu arkusza
    book = writer.book
    sheet = writer.sheets["Arkusz1"]

    # Ustawienie tytułu
    sheet.cell(row=1, column=startcol, value="Raport sprzedaży")
    sheet.cell(row=1, column=startcol).font = Font(size=24, bold=True)

    # Formatowanie arkusza
    sheet.sheet_view.showGridLines = False

    # Formatowanie obiektu DataFrame poprzez:
    # - format liczb
    # - szerokość kolumny
    # - formatowanie warunkowe
    for row in range(startrow + 1, startrow + nrows + 1):
        for col in range(startcol + 1, startcol + ncols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")

    for cell in sheet["B"]:
        cell.number_format = "mmm yy"

    for col in range(startcol, startcol + ncols + 1):
        cell = sheet.cell(row=startrow, column=col)
        sheet.column_dimensions[cell.column_letter].width = 14

    first_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
    last_cell = sheet.cell(row=startrow + nrows, column=startcol + ncols)
    range_address = f"{first_cell.coordinate}:{last_cell.coordinate}"
    sheet.conditional_formatting.add(range_address,
                                     CellIsRule(operator="lessThan",
                                                formula=["20000"],
                                                stopIfTrue=True,
                                                font=Font(color="E93423")))

    # Wykres
    chart = BarChart()
    chart.type = "col"
    chart.title = "Sprzedaż z podziałem na miesiące i sklepy"
    chart.height = 11.5
    chart.width = 20.5

    # Dodanie każdej kolumny jako serii, ignorując wiersz i kolumnę z podsumowaniem
    data = Reference(sheet, min_col=startcol + 1, min_row=startrow,
                     max_row=startrow + nrows - 1,
                     max_col=startcol + ncols - 1)
    categories = Reference(sheet, min_col=startcol, min_row=startrow + 1,
                           max_row=startrow + nrows - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    cell = sheet.cell(row=startrow + nrows + 2, column=startcol)
    sheet.add_chart(chart=chart, anchor=cell.coordinate)

    # Formatowanie wykresu
    chart.y_axis.title = "Sprzedaż"
    chart.x_axis.title = summary.index.name
    # Ukrycie linii osi y: spPR oznacza ShapeProperties 
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))