import multiprocessing
from itertools import repeat

import openpyxl
import excel


def _read_sheet(filename, sheetname):
    # Początkowy podkreślnik w nazwie funkcji jest zwyczajowo używany do 
	# oznaczenia jej jako "prywatnej", co oznacza, że nie powinna być używana
	# bezpośrednio poza tym modułem.
    book = openpyxl.load_workbook(filename,
                                  read_only=True, data_only=True)
    sheet = book[sheetname]
    data = excel.read(sheet)
    book.close()
    return sheet.title, data

def load_workbook(filename, sheetnames=None):
    if sheetnames is None:
        book = openpyxl.load_workbook(filename,
                                      read_only=True, data_only=True)
        sheetnames = book.sheetnames
        book.close()
    with multiprocessing.Pool() as pool:
        # Domyślnie Pool tworzy tyle procesów, ile jest rdzeni procesora.
		# starmap odwzorowuje krotkę argumentów na funkcję. Wyrażenie zip
		# tworzy listę zawierającą krotki o następującej postaci:
		# [('nazwa_pliku.xlsx', 'Arkusz1'), ('nazwa_pliku.xlsx', 'Arkusz2')]
        data = pool.starmap(_read_sheet, zip(repeat(filename), sheetnames))
    return {i[0]: i[1] for i in data}
